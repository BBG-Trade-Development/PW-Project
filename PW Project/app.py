import pandas as pd
import re
from datetime import datetime
import os
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from flask import Flask, request, render_template, send_file, flash, redirect, url_for, jsonify

#Test Comment

def resource_path(relative_path): 
    """Used for all related files (excel)"""
    base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
    return os.path.join(base_path, relative_path)

def get_output_dir():
    """Output folder created in same location as APP containing excel output"""
    base_path = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
    out_dir = os.path.join(base_path, 'outputs')
    os.makedirs(out_dir, exist_ok=True)
    return out_dir

app = Flask(__name__, template_folder=resource_path('templates'))
# app.secret_key = For when we use sharepoint/data access
output_dir = get_output_dir()

# Excel file -> use SharePoint
price_book_path = resource_path('Price_Book_Full.xlsx')
zpurcon_path = resource_path('ZPURCON.xlsx')

def improved_deduplication(df):
    dedup_strategies = [
        ['SAP Product ID', 'Deal ID', 'Deal Class', 'Channel', 'Purchase Quantity', 'Price Group', 'Start Date', 'End Date'],
        ['SAP Product ID', 'Deal ID', 'Deal Class', 'Channel', 'Purchase Quantity'],
        ['SAP Product ID', 'Price Group', 'Deal Class', 'Channel'],
        ['SAP Product ID', 'Price Group', 'Deal Class'],
        ['SAP Product ID', 'Price Group']
    ]
    
    for cols in dedup_strategies:
        existing_cols = [col for col in cols if col in df.columns]
        if len(existing_cols) >= 2 and df.duplicated(subset=existing_cols).any():
            return df.drop_duplicates(subset=existing_cols, keep='first')
    
    return df.copy()

def calculate_gp2_with_validation(df):
    required_cols = ['Case Price', 'Replacement Cost', 'Floor Stock Value', 'Chargeback']
    
    for col in required_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        else:
            df[col] = 0.0
    
    def calc_gp2(row, cost_col):
        case_price = row['Case Price']
        cost = row[cost_col]
        chargeback = row['Chargeback']
        if case_price != 0:
            cogs = cost - chargeback
            return (case_price - cogs) / case_price
        return pd.NA
    
    df['GP2 - Replacement Cost'] = df.apply(lambda row: calc_gp2(row, 'Replacement Cost'), axis=1)
    df['GP2 - Floor Stock Value'] = df.apply(lambda row: calc_gp2(row, 'Floor Stock Value'), axis=1)
    return df

def pivot_key_sort_key(pivot_key_str):
    try:
        parts = pivot_key_str.split(' | ')
        if len(parts) != 4:
            print(f"Invalid Pivot Key: {pivot_key_str}")
            return (99, 99, 99, 9999)
        channel = parts[0]
        pricing_type = parts[1]
        deal_class = parts[2]
        purchase_qty = parts[3]

        channel_order = {'Retail': 0, 'OP': 1}
        channel_val = channel_order.get(channel, 99)
        if channel_val == 99:
            print(f"Unknown Channel in Pivot Key: {channel}")
        
        pricing_order = {'Level Pricing': 0, 'Deal Pricing': 1}
        pricing_val = pricing_order.get(pricing_type, 99)
        
        deal_class_order = [
            'Level Pricing', 'EVD – Straight Discount', 'Close – Straight Discount',
            'Promo – Straight Discount', 'EVD- Special Price Goods', 'Promo- Special Price Goods',
            'Inventory Reduction – Straight Discount', 'Inventory Reduction- Special Price Goods'
        ]
        try:
            deal_class_val = deal_class_order.index(deal_class)
        except ValueError:
            deal_class_val = 99
        
        qty_match = re.match(r'(\d+)', purchase_qty)
        qty_val = int(qty_match.group(1)) if qty_match else 9999
        
        return (channel_val, pricing_val, deal_class_val, qty_val)
    except Exception as e:
        print(f"Error in pivot_key_sort_key: {pivot_key_str}, Exception: {str(e)}")
        return (99, 99, 99, 9999)

def process_data(vendor_id, gp2_threshold):
    try:
        PB = pd.read_excel(price_book_path, sheet_name='Printer Friendly')
        ZPUR = pd.read_excel(zpurcon_path)
    except FileNotFoundError as e:
        return None, f"Excel file not found: {e}"
    except Exception as e:
        return None, f"Error loading Excel files: {e}"

    cols_to_merge = [
        'Supplier', 'Price Group #', 'Price Group Description', 'FOB', 'SPA', 'Miscellaneous',
        'Land Freight', 'Ocean Freight', 'Federal Tax', 'Broker Charge', 'Bulk Whiskey Fee',
        'Duty', 'Tariffs Per Case', 'Consolidate Fee', 'Gallonage tax per case pd to Vendor',
        'Gallonage tax per case Pd to State', 'Gallonage tax Volume based Pd to State',
        'Total', 'Mov Avg 7210', 'Stock in bottles', 'Stock in Cases', 'Mrp Controller'
    ]
    existing_cols_to_merge = [col for col in cols_to_merge if col in ZPUR.columns]
    if not existing_cols_to_merge:
        return None, "No matching columns found for merging data."

    PB_merged = PB.merge(
        ZPUR[['Material'] + existing_cols_to_merge],
        left_on='SAP Product ID',
        right_on='Material',
        how='left'
    ).drop(columns=['Material'])

    if 'Supplier' not in PB_merged.columns:
        return None, "Supplier column missing in merged data."
    PB_merged['Supplier'] = pd.Series(PB_merged['Supplier'], dtype='object').apply(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace('.', '').isdigit() else '0'
    ).str.zfill(6)
    PB_input = PB_merged[PB_merged['Supplier'] == vendor_id].copy()
    if PB_input.empty:
        return None, f"Vendor ID {vendor_id} not found."

    # Determine vendor name
    vendor_name = "UnknownVendor" #If not found
    if 'Vendor' in PB_input.columns and not PB_input['Vendor'].dropna().empty:
        vendor_name = str(PB_input['Vendor'].dropna().iloc[0])
    else:
        
        vendor_name = f"Vendor_{vendor_id}"
    vendor_name_sanitized = re.sub(r'[^\w-]', '', vendor_name).replace(' ', '_').strip('_') or "NoName"

    # === COGS Processing ===
    cols_to_drop_cogs = [
        'Vendor', 'UPC Bottle', 'UPC Cases', 'UPC Sleeve', 'Business Manager Detail', 'Pricing Type',
        'Deal ID', 'Deal Class', 'Trade Channel ID', 'Chain Name', 'Purchase Quantity', 'Deal Description',
        'Start Date', 'End Date', 'Discount', 'Chargeback', 'Chargeback Type', 'Case Price', 'Bottle Price',
        'Units Per Case', 'Supplier', 'Price Group #', 'Price Group Description_y', 'Broker Charge',
        'Bulk Whiskey Fee', 'Consolidate Fee', 'Mrp Controller'
    ]
    cogs = PB_input.drop(columns=[col for col in cols_to_drop_cogs if col in PB_input.columns]).drop(columns=['Price Group Description_y'], errors='ignore').drop_duplicates(subset=['SAP Product ID'])
    cogs = cogs.rename(columns={
        'Miscellaneous': 'Misc', 'Land Freight': 'Land Frt', 'Ocean Freight': 'Ocean Frt',
        'Federal Tax': 'Fed Tax', 'Tariffs Per Case': 'Tariff', 'Gallonage tax per case pd to Vendor': 'Tax pd to Ven',
        'Gallonage tax per case Pd to State': 'State Tax Cs', 'Gallonage tax Volume based Pd to State': 'State Tax Vol',
        'Mov Avg 7210': 'Floor Stock Value', 'Stock in bottles': 'Btls OH', 'Stock in Cases': 'Cases OH',
        'Price Group Description_x': 'Price Group Description', 'Total': 'Replacement Cost'
    })
    if 'Size' in cogs.columns:
        cogs = cogs[~cogs['Size'].str.startswith('COMBO', na=False)]
    if 'List Case' in cogs.columns and 'List Price' not in cogs.columns:
        cogs['List Price'] = cogs['List Case']

    # === PW Processing ===
    cols_to_drop_PW = [
        'Vendor', 'Group Name', 'Size', 'Product Name', 'SAP Product ID', 'UPC Bottle', 'UPC Cases', 'UPC Sleeve',
        'Business Manager Detail', 'Chain Name', 'Supplier', 'Price Group #', 'Price Group Description_y', 'FOB', 'SPA',
        'Miscellaneous', 'Land Freight', 'Ocean Freight', 'Federal Tax', 'Broker Charge', 'Bulk Whiskey Fee', 'Duty',
        'Tariffs Per Case', 'Consolidate Fee', 'Gallonage tax per case pd to Vendor', 'Gallonage tax per case Pd to State',
        'Gallonage tax Volume based Pd to State', 'Stock in bottles', 'Mrp Controller'
    ]
    PW = PB_input.drop(columns=[col for col in cols_to_drop_PW if col in PB_input.columns], errors='ignore')
    
    PW = PW.dropna(subset=['Price Group'])
    
    values_to_exclude_pricing_type = ['Volume Incentives', 'Chain Pricing']
    PW = PW[
        PW['Pricing Type'].notna() &
        (PW['Pricing Type'].astype(str).str.strip() != '') &
        (~PW['Pricing Type'].isin(values_to_exclude_pricing_type))
    ]
    
    values_to_exclude_TC = ['C2','C3','C4','C5','C6','C7','C8','C9','C10','C11','C12','C13','C14','C15','C17','C18','C19']
    PW = PW[~PW['Trade Channel ID'].isin(values_to_exclude_TC)]
    
    values_to_exclude_PQ = ['0', '0 CSE', '0 EA']
    PW['Purchase Quantity Clean'] = PW['Purchase Quantity'].astype(str).str.strip()
    PW = PW[
        (~PW['Purchase Quantity Clean'].isin(values_to_exclude_PQ)) &
        (PW['Purchase Quantity Clean'] != '')
    ]
    PW = PW.drop(columns=['Purchase Quantity Clean'])
    
    PW = PW.rename(columns={'Price Group Description_x': 'Price Group Description'})
    PW = PW.rename(columns={'Trade Channel ID': 'Channel'})
    PW['Channel'] = PW['Channel'].astype(str).str.strip()
    PW['Channel'] = PW['Channel'].replace({'C1': 'Retail', 'C16': 'OP', '': 'OP', 'nan': 'OP'})
    
    for date_col in ['Start Date', 'End Date']:
        if date_col in PW.columns:
            PW[date_col] = pd.to_datetime(PW[date_col], errors='coerce').dt.date
    
    PW = PW.rename(columns={
        'List Price': 'List Case',
        'Total': 'Replacement Cost',
        'Mov Avg 7210': 'Floor Stock Value',
        'Stock in Cases': 'Cases OH'
    })
    
    if 'Cases OH' in PW.columns and 'Floor Stock Value' in PW.columns:
        PW['FSV_Weighted_Numerator_Temp'] = PW['Floor Stock Value'] * PW['Cases OH']
        PW['FSV_Weighted_Denominator_Temp'] = PW['Cases OH']
        weighted_avg_fsv_global = PW.groupby('Price Group').apply(
            lambda x: x['FSV_Weighted_Numerator_Temp'].sum() / x['FSV_Weighted_Denominator_Temp'].sum()
            if x['FSV_Weighted_Denominator_Temp'].sum() != 0 else x['Floor Stock Value'].iloc[0] if not x['Floor Stock Value'].empty else 0
        ).rename('Weighted_Avg_Floor_Stock_Value_Global')
        PW = PW.merge(weighted_avg_fsv_global.reset_index(), on='Price Group', how='left')
        PW['Floor Stock Value'] = PW['Weighted_Avg_Floor_Stock_Value_Global']
        PW = PW.drop(columns=['FSV_Weighted_Numerator_Temp', 'FSV_Weighted_Denominator_Temp', 'Weighted_Avg_Floor_Stock_Value_Global'], errors='ignore')
    
    PW = PW[~PW['Price Group Description'].str.startswith('COMBO', na=False)]
    
    PW_deduped = improved_deduplication(PW).copy()

    filename = f"PW_{vendor_id}_{vendor_name_sanitized}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(output_dir, filename)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        percentage_format = '0.00%'
        light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        medium_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        cogs_currency_cols = [
            'List Price', 'FOB', 'SPA', 'Misc', 'Land Frt', 'Ocean Frt', 'Fed Tax', 'Duty', 'Tariff',
            'Tax pd to Ven', 'State Tax Cs', 'State Tax Vol', 'Replacement Cost', 'Floor Stock Value'
        ]
        pw_currency_cols = [
            'List Case', 'List Bottle', 'Discount', 'Case Price', 'Bottle Price', 'Chargeback',
            'Replacement Cost', 'Floor Stock Value'
        ]
        pw_percentage_cols = ['GP2 - Replacement Cost', 'GP2 - Floor Stock Value']

        cogs.to_excel(writer, index=False, sheet_name='COGS')
        worksheet = writer.sheets['COGS']
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = light_blue_fill
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 4
        worksheet.freeze_panes = 'A2'
        header_map_cogs = {cell.value: cell.column for cell in worksheet[1]}
        for col_name in cogs_currency_cols:
            if col_name in header_map_cogs:
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=header_map_cogs[col_name])
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = accounting_format

        if 'Price Group' in cogs.columns and 'Replacement Cost' in cogs.columns:
            price_group_consistency = cogs.groupby('Price Group')['Replacement Cost'].nunique()
            error_price_groups = price_group_consistency[price_group_consistency > 1].index.tolist()
            if error_price_groups:
                cogs_errors_df = cogs[cogs['Price Group'].isin(error_price_groups)].sort_values(['Price Group', 'SAP Product ID'])
                if 'Price Group Description_y' in cogs_errors_df.columns:
                    cogs_errors_df = cogs_errors_df.drop(columns=['Price Group Description_y'])
                cogs_errors_df.to_excel(writer, index=False, sheet_name='Price Group Errors')
                worksheet_errors = writer.sheets['Price Group Errors']
                worksheet_errors.sheet_properties.tabColor = "FF9999"
                for col_idx in range(1, worksheet_errors.max_column + 1):
                    cell = worksheet_errors.cell(row=1, column=col_idx)
                    cell.fill = light_blue_fill
                    cell.font = Font(bold=True)
                for column_cells in worksheet_errors.columns:
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                    worksheet_errors.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 4
                worksheet_errors.freeze_panes = 'A2'
                header_map_errors = {cell.value: cell.column for cell in worksheet_errors[1]}
                for col_name in cogs_currency_cols:
                    if col_name in header_map_errors:
                        for row_idx in range(2, worksheet_errors.max_row + 1):
                            cell = worksheet_errors.cell(row=row_idx, column=header_map_errors[col_name])
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = accounting_format
                            if col_name == 'Replacement Cost':
                                cell.fill = light_red_fill

        # === GP2 Below Threshold Tab ===
        gp2_filtered_df = PW_deduped.copy()
        for col in ['Units Per Case', 'Case Price', 'Replacement Cost', 'Chargeback', 'List Case']:
            if col in gp2_filtered_df.columns:
                gp2_filtered_df[col] = pd.to_numeric(gp2_filtered_df[col], errors='coerce').fillna(0)
        gp2_filtered_df['List Bottle'] = gp2_filtered_df.apply(
            lambda row: row['List Case'] / row['Units Per Case'] if row['Units Per Case'] != 0 else pd.NA, axis=1
        )
        gp2_filtered_df['Bottle Price'] = gp2_filtered_df.apply(
            lambda row: row['Case Price'] / row['Units Per Case'] if row['Units Per Case'] != 0 else pd.NA, axis=1
        )
        gp2_filtered_df = calculate_gp2_with_validation(gp2_filtered_df)
        gp2_below_threshold = gp2_filtered_df[
            ((gp2_filtered_df['GP2 - Replacement Cost'].notna()) & (gp2_filtered_df['GP2 - Replacement Cost'] < gp2_threshold)) |
            ((gp2_filtered_df['GP2 - Floor Stock Value'].notna()) & (gp2_filtered_df['GP2 - Floor Stock Value'] < gp2_threshold))
        ].copy()
        
        display_cols = [
            'Price Group', 'Price Group Description', 'Pricing Type', 'Deal ID', 'Deal Class',
            'Purchase Quantity', 'Deal Description', 'Start Date', 'End Date', 'List Case',
            'Discount', 'Chargeback', 'Replacement Cost', 'Floor Stock Value', 'List Bottle',
            'Case Price', 'Bottle Price', 'GP2 - Replacement Cost', 'GP2 - Floor Stock Value'
        ]
        existing_display_cols = [col for col in display_cols if col in gp2_below_threshold.columns]
        gp2_output_df = gp2_below_threshold[existing_display_cols].copy() if not gp2_below_threshold.empty else pd.DataFrame(columns=existing_display_cols)

        gp2_output_df.to_excel(writer, index=False, sheet_name='GP2 Below Threshold')
        worksheet_gp2 = writer.sheets['GP2 Below Threshold']
        worksheet_gp2.sheet_properties.tabColor = "FF9999"
        for col_idx in range(1, worksheet_gp2.max_column + 1):
            cell = worksheet_gp2.cell(row=1, column=col_idx)
            cell.fill = light_blue_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for column_cells in worksheet_gp2.columns:
            column_letter = get_column_letter(column_cells[0].column)
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            if column_letter in ['M', 'N', 'O', 'P', 'Q']:
                worksheet_gp2.column_dimensions[column_letter].width = max(length + 6, 15)
            else:
                worksheet_gp2.column_dimensions[column_letter].width = length + 4
        worksheet_gp2.freeze_panes = 'A2'
        if not gp2_output_df.empty:
            header_map_gp2 = {cell.value: cell.column for cell in worksheet_gp2[1]}
            for col_name in pw_currency_cols:
                if col_name in header_map_gp2:
                    for row_idx in range(2, worksheet_gp2.max_row + 1):
                        cell = worksheet_gp2.cell(row=row_idx, column=header_map_gp2[col_name])
                        if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.number_format = accounting_format
            for col_name in pw_percentage_cols:
                if col_name in header_map_gp2:
                    for row_idx in range(2, worksheet_gp2.max_row + 1):
                        cell = worksheet_gp2.cell(row=row_idx, column=header_map_gp2[col_name])
                        if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.number_format = percentage_format
                            if cell.value < gp2_threshold:
                                cell.fill = light_red_fill
            for date_col in ['Start Date', 'End Date']:
                if date_col in header_map_gp2:
                    for row_idx in range(2, worksheet_gp2.max_row + 1):
                        cell = worksheet_gp2.cell(row=row_idx, column=header_map_gp2[date_col])
                        if isinstance(cell.value, (datetime, pd.Timestamp)) and not pd.isna(cell.value):
                            cell.number_format = 'MM/DD/YYYY'
        else:
            worksheet_gp2.cell(row=2, column=1, value=f"No records found with GP2 margins below {gp2_threshold:.1%}")
            worksheet_gp2.cell(row=2, column=1).font = Font(italic=True, color="666666")

        # === Brand Tabs ===
        if 'Brand' in PW_deduped.columns:
            print(f"Brand column found. Unique brands: {PW_deduped['Brand'].dropna().unique()}")
            if PW_deduped['Brand'].dropna().empty:
                print("Warning: No non-NaN Brand values found in PW_deduped.")
                pd.DataFrame({"Message": ["No valid Brand data found"]}).to_excel(writer, index=False, sheet_name='No_Brands')
                worksheet_no_brands = writer.sheets['No_Brands']
                worksheet_no_brands.cell(row=1, column=1).font = Font(italic=True, color="666666")
            else:
                for brand in PW_deduped['Brand'].dropna().unique():
                    print(f"Processing brand: {brand}")
                    brand_df = PW_deduped[PW_deduped['Brand'] == brand].copy()
                    if brand_df.empty:
                        print(f"Warning: No data for brand {brand} after filtering.")
                        continue

                    # Perform calculations
                    for col in ['Units Per Case', 'Case Price', 'Replacement Cost', 'Chargeback', 'List Case']:
                        if col in brand_df.columns:
                            brand_df[col] = pd.to_numeric(brand_df[col], errors='coerce').fillna(0)
                    brand_df['List Bottle'] = brand_df.apply(
                        lambda row: row['List Case'] / row['Units Per Case'] if row['Units Per Case'] != 0 else pd.NA, axis=1)
                    brand_df['Bottle Price'] = brand_df.apply(
                        lambda row: row['Case Price'] / row['Units Per Case'] if row['Units Per Case'] != 0 else pd.NA, axis=1)
                    brand_df['GP2 - Replacement Cost'] = brand_df.apply(
                        lambda row: (row['Case Price'] - (row.get('Replacement Cost', 0) - row.get('Chargeback', 0))) / row['Case Price']
                        if row.get('Case Price', 0) != 0 else pd.NA, axis=1)
                    brand_df['GP2 - Floor Stock Value'] = brand_df.apply(
                        lambda row: (row['Case Price'] - (row.get('Floor Stock Value', 0) - row.get('Chargeback', 0))) / row['Case Price']
                        if row.get('Case Price', 0) != 0 else pd.NA, axis=1)

                    pivot_key_cols = ['Channel', 'Pricing Type', 'Deal Class', 'Purchase Quantity']
                    existing_pivot_cols = [col for col in pivot_key_cols if col in brand_df.columns]
                    print(f"Brand: {brand}, Available pivot columns: {existing_pivot_cols}")

                    if len(existing_pivot_cols) == 4:
                        # Ensure Channel values are clean
                        brand_df['Channel'] = brand_df['Channel'].astype(str).str.strip()
                        brand_df['Channel'] = brand_df['Channel'].replace({'C1': 'Retail', 'C16': 'OP', '': 'OP', 'nan': 'OP'})
                        
                        # Create Pivot Key and validate
                        brand_df['Pivot Key'] = brand_df[existing_pivot_cols].astype(str).agg(' | '.join, axis=1)
                        
                        # Debug: Log unique Pivot Key values
                        print(f"Brand: {brand}, Unique Pivot Keys: {brand_df['Pivot Key'].unique()}")
                        
                        # Validate Pivot Key format
                        brand_df['Pivot Key Valid'] = brand_df['Pivot Key'].apply(
                            lambda x: len(x.split(' | ')) == 4 and x.split(' | ')[0] in ['Retail', 'OP']
                        )
                        if not brand_df['Pivot Key Valid'].all():
                            print(f"Warning: Invalid Pivot Keys found for brand {brand}:")
                            print(brand_df[~brand_df['Pivot Key Valid']]['Pivot Key'].unique())
                        
                        id_vars = [col for col in ['Vendor ID', 'Brand', 'Price Group', 'Price Group Description', 'Units Per Case', 'Pivot Key'] if col in brand_df.columns]
                        value_vars = [col for col in [
                            'Replacement Cost', 'Floor Stock Value', 'List Case', 'List Bottle', 'Discount', 'Case Price',
                            'Bottle Price', 'Chargeback', 'GP2 - Replacement Cost', 'GP2 - Floor Stock Value',
                            'Start Date', 'End Date', 'Deal ID', 'Deal Description'
                        ] if col in brand_df.columns]
                        try:
                            brand_melted = brand_df.melt(id_vars=id_vars, value_vars=value_vars, var_name='Product Cost Breakdown', value_name='Value')
                            brand_pivot = brand_melted.pivot_table(
                                index=[col for col in ['Brand', 'Price Group', 'Price Group Description', 'Units Per Case', 'Product Cost Breakdown'] if col in brand_melted.columns],
                                columns='Pivot Key', values='Value', aggfunc='first').reset_index().drop(columns=['Brand'], errors='ignore')
                            desired_order = [col for col in value_vars if col in brand_pivot['Product Cost Breakdown'].values]
                            if 'Product Cost Breakdown' in brand_pivot.columns:
                                brand_pivot['Product Cost Breakdown'] = pd.Categorical(brand_pivot['Product Cost Breakdown'], categories=desired_order, ordered=True)
                                brand_pivot = brand_pivot.sort_values(['Price Group', 'Price Group Description', 'Units Per Case', 'Product Cost Breakdown'])
                                group_cols = ['Price Group', 'Price Group Description', 'Units Per Case']
                                if group_cols[0] in brand_pivot.columns:
                                    block_change = brand_pivot[group_cols].ne(brand_pivot[group_cols].shift()).any(axis=1)
                                    for col in group_cols:
                                        if col in brand_pivot.columns:
                                            brand_pivot[col] = brand_pivot[col].where(block_change, '')
                            
                            # Sort Pivot Key columns
                            pivot_cols = [col for col in brand_pivot.columns if col not in ['Price Group', 'Price Group Description', 'Units Per Case', 'Product Cost Breakdown']]
                            print(f"Brand: {brand}, Pivot Columns Before Sorting: {pivot_cols}")
                            sorted_pivot_cols = sorted(pivot_cols, key=pivot_key_sort_key)
                            print(f"Brand: {brand}, Pivot Columns After Sorting: {sorted_pivot_cols}")
                            brand_pivot = brand_pivot[['Price Group', 'Price Group Description', 'Units Per Case', 'Product Cost Breakdown'] + sorted_pivot_cols]
                        except Exception as e:
                            print(f"Error creating pivot table for brand {brand}: {str(e)}")
                            # Fallback to simple table if pivot fails
                            brand_cols = [col for col in display_cols if col in brand_df.columns]
                            brand_simple = brand_df[brand_cols].copy()
                            safe_brand_name = str(brand)[:31]
                            brand_simple.to_excel(writer, index=False, sheet_name=safe_brand_name)
                            print(f"Fallback: Wrote simple table for brand {brand}")
                            continue
                    else:
                        print(f"Brand: {brand}, Insufficient columns for pivot table: {existing_pivot_cols}")
                        brand_cols = [col for col in display_cols if col in brand_df.columns]
                        brand_simple = brand_df[brand_cols].copy()
                        safe_brand_name = str(brand)[:31]
                        brand_simple.to_excel(writer, index=False, sheet_name=safe_brand_name)
                        print(f"Wrote simple table for brand {brand}")
                        continue

                    safe_brand_name = str(brand)[:31]
                    try:
                        brand_pivot.to_excel(writer, index=False, sheet_name=safe_brand_name)
                        print(f"Successfully wrote pivot table for brand {brand}")
                    except Exception as e:
                        print(f"Error writing pivot table for brand {brand}: {str(e)}")
                        continue

                    worksheet_brand = writer.sheets[safe_brand_name]
                    # Insert 3 rows at the top for stacked headers
                    worksheet_brand.insert_rows(1, amount=3)
                    # Apply light blue background and bold font to header rows
                    for r_idx in range(1, 5):
                        for c_idx in range(1, worksheet_brand.max_column + 1):
                            cell = worksheet_brand.cell(row=r_idx, column=c_idx)
                            cell.fill = light_blue_fill
                            cell.font = Font(bold=True)
                    # Set custom headers for row 4
                    final_row_4_headers_map = {
                        'Price Group': 'Price Group',
                        'Price Group Description': 'Price Group Description',
                        'Units Per Case': 'PK',
                        'Product Cost Breakdown': 'Pricing Details'
                    }
                    for col_idx in range(1, worksheet_brand.max_column + 1):
                        cell_in_row4 = worksheet_brand.cell(row=4, column=col_idx)
                        original_header = cell_in_row4.value
                        if original_header in final_row_4_headers_map:
                            target_cell = worksheet_brand.cell(row=4, column=col_idx)
                            target_cell.value = final_row_4_headers_map[original_header]
                            target_cell.alignment = Alignment(horizontal='center')
                            for row_offset in range(3):
                                worksheet_brand.cell(row=1 + row_offset, column=col_idx, value='')
                        elif original_header and isinstance(original_header, str) and ' | ' in original_header:
                            parts = original_header.split(' | ')
                            while len(parts) < 4:
                                parts.append('')
                            header_parts_to_write = [parts[0], parts[1], parts[2], parts[3]]
                            for row_offset, part_value in enumerate(header_parts_to_write):
                                header_cell = worksheet_brand.cell(row=1 + row_offset, column=col_idx, value=part_value)
                                header_cell.alignment = Alignment(horizontal='center')
                        else:
                            target_cell = worksheet_brand.cell(row=4, column=col_idx)
                            target_cell.alignment = Alignment(horizontal='center')
                            for row_offset in range(3):
                                worksheet_brand.cell(row=1 + row_offset, column=col_idx, value='')
                    # Apply bold formatting to columns A:D
                    for r_idx in range(1, worksheet_brand.max_row + 1):
                        for c_idx in range(1, 5):
                            cell = worksheet_brand.cell(row=r_idx, column=c_idx)
                            cell.font = Font(bold=True)
                    
                    # Freeze panes at E5
                    worksheet_brand.freeze_panes = 'E5'
                    # Auto-size columns
                    for column_cells in worksheet_brand.columns:
                        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                        column_letter = get_column_letter(column_cells[0].column)
                        worksheet_brand.column_dimensions[column_letter].width = length + 4

                    # Apply borders
                    side_medium = Side(style='medium')
                    pcb_idx = brand_pivot.columns.get_loc('Product Cost Breakdown') + 1
                    pivot_key_excel_col_indices = []
                    start_col_for_pivot_keys = pcb_idx + 1
                    for col_idx_openpyxl in range(start_col_for_pivot_keys, worksheet_brand.max_column + 1):
                        pivot_key_excel_col_indices.append(col_idx_openpyxl)
                    
                    rows_for_bottom_border = []
                    for r_idx_check in range(5, worksheet_brand.max_row + 1):
                        pcb_cell_value = worksheet_brand.cell(row=r_idx_check, column=pcb_idx).value
                        if pcb_cell_value == 'Deal Description':
                            rows_for_bottom_border.append(r_idx_check)
                    
                    for c_idx in range(1, worksheet_brand.max_column + 1):
                        cell = worksheet_brand.cell(row=4, column=c_idx)
                        current_border = cell.border if cell.border else Border()
                        cell.border = Border(
                            left=current_border.left, right=current_border.right,
                            top=current_border.top, bottom=side_medium
                        )

                    for r_idx in range(1, worksheet_brand.max_row + 1):
                        for c_idx in range(1, worksheet_brand.max_column + 1):
                            cell = worksheet_brand.cell(row=r_idx, column=c_idx)
                            current_border = cell.border if cell.border else Border()

                            new_left = current_border.left
                            new_right = current_border.right
                            new_top = current_border.top
                            new_bottom = current_border.bottom

                            if r_idx in rows_for_bottom_border:
                                new_bottom = side_medium
                            if c_idx == pcb_idx:
                                new_right = side_medium
                            if c_idx in pivot_key_excel_col_indices:
                                if c_idx == pivot_key_excel_col_indices[-1] or (c_idx + 1) not in pivot_key_excel_col_indices:
                                    new_right = side_medium
                            
                            cell.border = Border(left=new_left, right=new_right, top=new_top, bottom=new_bottom)
        else:
            print("No Brand column found in PW_deduped.")
            pd.DataFrame({"Message": ["No Brand column available"]}).to_excel(writer, index=False, sheet_name='No_Brands')
            worksheet_no_brands = writer.sheets['No_Brands']
            worksheet_no_brands.cell(row=1, column=1).font = Font(italic=True, color="666666")

        return filename, None

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        vendor_id = request.form.get('vendor_id', '').strip()
        gp2_threshold = request.form.get('gp2_threshold', '').strip()
        if not (vendor_id.isdigit() and len(vendor_id) == 6 and vendor_id.startswith('3')):
            flash("Invalid Vendor ID.", 'error')
            return redirect(url_for('index'))
        try:
            gp2_threshold_val = float(gp2_threshold)
            if not (0 <= gp2_threshold_val <= 1):
                raise ValueError()
        except ValueError:
            flash("Invalid GP2 threshold.", 'error')
            return redirect(url_for('index'))
        filename, error = process_data(vendor_id, gp2_threshold_val)
        if error:
            flash(error, 'error')
            return redirect(url_for('index'))
        flash(f"Processing complete! Download: {filename}", 'success')
        return redirect(url_for('download_file', filename=filename))
    return render_template('index.html')

@app.route('/api/process', methods=['POST'])
def api_process():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No JSON data received'}), 400
        vendor_id = str(data.get('vendor_id', '')).strip()
        gp2_threshold = str(data.get('gp2_threshold', '')).strip()
        if not (vendor_id.isdigit() and len(vendor_id) == 6 and vendor_id.startswith('3')):
            return jsonify({'success': False, 'message': "Invalid Vendor ID."}), 400
        try:
            gp2_threshold_val = float(gp2_threshold)
            if not (0 <= gp2_threshold_val <= 1):
                raise ValueError()
        except ValueError:
            return jsonify({'success': False, 'message': "Invalid GP2 threshold."}), 400
        filename, error = process_data(vendor_id, gp2_threshold_val)
        if error:
            return jsonify({'success': False, 'message': error}), 500
        return jsonify({
            'success': True,
            'message': f'File generated: {filename}',
            'download_url': url_for('download_file', filename=filename, _external=True)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    path = os.path.join(output_dir, filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name=filename)
    flash("File not found.", 'error')
    return redirect(url_for('index'))

def open_browser():
    import webbrowser
    webbrowser.open_new('http://127.0.0.1:5000')

if __name__ == '__main__':
    import threading
    threading.Timer(1.0, open_browser).start()
    app.run(debug=False, host='127.0.0.1', port=5000)