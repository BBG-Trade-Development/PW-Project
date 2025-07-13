
import os
import re
import shutil
import sys
import tempfile
import threading
import webbrowser
from datetime import datetime

import numpy as np
import pandas as pd
from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def strftime_filter(value, format_string="%Y-%m-%d"):
    """Format a datetime object or string to a specified format."""
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value
    if isinstance(value, datetime):
        return value.strftime(format_string)
    return str(value)

def resource_path(relative_path):
    """Get path for non-Excel files (e.g., templates)."""
    base_path = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base_path, relative_path)

def get_output_dir():
    """Create output folder in same location as app for Excel output."""
    base_path = os.path.dirname(
        sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
    )
    out_dir = os.path.join(base_path, "outputs")
    os.makedirs(out_dir, exist_ok=True)
    return out_dir

app = Flask(__name__, template_folder=resource_path("templates"))
app.secret_key = os.urandom(24)
app.jinja_env.filters["strftime"] = strftime_filter
output_dir = get_output_dir()

def get_sharepoint_context(username, password):
    """Authenticate with SharePoint using user credentials."""
    sharepoint_url = "https://skyappscsg.sharepoint.com/teams/TDAnalysts-BBGCA"
    if not username or not password:
        return None
    try:
        credentials = UserCredential(username, password)
        ctx = ClientContext(sharepoint_url).with_credentials(credentials)
        web = ctx.web.get().execute_query()
        print(f"Connected to SharePoint site: {web.properties['Title']}")
        return ctx
    except Exception as e:
        print(f"Error connecting to SharePoint: {str(e)}")
        if "401" in str(e) or "Unauthorized" in str(e):
            print("Authentication failed: Check email and password.")
        elif "403" in str(e) or "Forbidden" in str(e):
            print("Permission denied: Check account permissions for the site.")
        return None

def download_sharepoint_file(ctx, relative_path, temp_dir):
    """Download a file from SharePoint to a temporary directory."""
    try:
        file_path = os.path.join(temp_dir, os.path.basename(relative_path))
        server_relative_url = f"/teams/TDAnalysts-BBGCA/Shared Documents/{relative_path}"
        with open(file_path, "wb") as local_file:
            file = ctx.web.get_file_by_server_relative_url(server_relative_url)
            file.download(local_file).execute_query()
        print(f"Downloaded {relative_path} to {file_path}")
        return file_path
    except Exception as e:
        error_message = f"Error downloading {relative_path} from {server_relative_url}: {str(e)}"
        print(error_message)
        if "401" in str(e) or "Unauthorized" in str(e):
            print("Authentication error: Check email and password.")
        elif "404" in str(e) or "FileNotFound" in str(e):
            print(f"File not found at {server_relative_url}. Verify file path and name.")
        elif "403" in str(e) or "Forbidden" in str(e):
            print(f"Permission denied for {server_relative_url}. Check account permissions.")
        else:
            print(f"Unexpected error: {str(e)}")
        return None

def improved_deduplication(df):
    """Deduplicate DataFrame based on specified column combinations."""
    dedup_strategies = [
        [
            "SAP Product ID",
            "Deal ID",
            "Deal Class",
            "Channel",
            "Purchase Quantity",
            "Price Group",
            "Start Date",
            "End Date",
        ],
        ["SAP Product ID", "Deal ID", "Deal Class", "Channel", "Purchase Quantity"],
        ["SAP Product ID", "Price Group", "Deal Class", "Channel"],
        ["SAP Product ID", "Price Group", "Deal Class"],
        ["SAP Product ID", "Price Group"],
    ]
    for cols in dedup_strategies:
        existing_cols = [col for col in cols if col in df.columns]
        if len(existing_cols) >= 2 and df.duplicated(subset=existing_cols).any():
            return df.drop_duplicates(subset=existing_cols, keep="first")
    return df.copy()

def calculate_gp2_with_validation(df, skip_gp2_if_no_price=False, price_col="Case Price"):
    """Calculate GP2 metrics with validation, using specified price column."""
    required_cols = [price_col, "Negotiated Cost", "Avg Cost", "Chargeback"]
    for col in required_cols:
        if col in df.columns:
            non_numeric = df[col].apply(lambda x: not isinstance(x, (int, float)) and pd.notna(x))
            if non_numeric.any():
                print(f"Warning: Non-numeric values in {col}:\n{df[non_numeric][['Sap Product Id', col]].head().to_string()}")
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            print(f"Warning: Column {col} missing in DataFrame. Setting to 0.")
            df[col] = 0.0
    if skip_gp2_if_no_price and (price_col not in df.columns or df[price_col].eq(0).all()):
        print(f"Skipping GP2 calculations due to missing or invalid {price_col}.")
        return df
    def calc_gp2(row, cost_col):
        try:
            price = row[price_col]
            cost = row[cost_col]
            chargeback = row["Chargeback"]
            if price != 0:
                cogs = cost - chargeback
                gp2 = (price - cogs) / price
                if not isinstance(gp2, (int, float)) or pd.isna(gp2):
                    print(
                        f"Invalid GP2 for {cost_col} at Sap Product Id {row.get('Sap Product Id', 'Unknown')}: "
                        f"{price_col.lower()}={price}, cost={cost}, chargeback={chargeback}"
                    )
                    return pd.NA
                return gp2
            return pd.NA
        except Exception as e:
            print(
                f"Error in GP2 for {cost_col} at Sap Product Id {row.get('Sap Product Id', 'Unknown')}: {str(e)}"
            )
            return pd.NA
    try:
        df["GP2 - Negotiated Cost"] = df.apply(lambda row: calc_gp2(row, "Negotiated Cost"), axis=1)
        df["GP2 - Avg Cost"] = df.apply(lambda row: calc_gp2(row, "Avg Cost"), axis=1)
        print(
            f"GP2 calculations completed. Non-null GP2 - Negotiated Cost: {df['GP2 - Negotiated Cost'].notna().sum()}, "
            f"GP2 - Avg Cost: {df['GP2 - Avg Cost'].notna().sum()}"
        )
    except Exception as e:
        print(f"Error in calculate_gp2_with_validation: {str(e)}")
        raise ValueError(f"Error processing {price_col} in GP2 calculation: {str(e)}")
    return df

def process_data(vendor_id, gp2_threshold, username, password, date_entry):
    """Process data and generate Excel output with pricing information."""
    temp_dir = tempfile.mkdtemp()
    try:
        # Initialize chain_output_df with default empty DataFrame to prevent undefined variable errors
        chain_output_df = pd.DataFrame(
            columns=[
                "Vendor Id",
                "Price Group",
                "Price Group Description",
                "Chain Name",
                "Start Date",
                "End Date",
                "List Price",
                "Net Price",
                "Bottle Price",
                "Chargeback",
                "Negotiated Cost",
                "Avg Cost",
                "GP2 - Negotiated Cost",
                "GP2 - Avg Cost",
            ]
        )
        print("Initialized chain_output_df with empty DataFrame")
        
        try:
            date_entry = datetime.strptime(date_entry, "%Y-%m-%d").date()
        except ValueError:
            return None, "Invalid date format. Use YYYY-MM-DD."
        ctx = get_sharepoint_context(username, password)
        if not ctx:
            return None, "Failed to connect to SharePoint. Check email and password."
        price_book_relative_path = "PW Project/Price_Book_Full.xlsx"
        zpurcon_relative_path = "PW Project/ZPURCON.xlsx"
        chain_pricing_relative_path = "PW Project/Chain_Pricing.xlsx"
        price_book_path = download_sharepoint_file(ctx, price_book_relative_path, temp_dir)
        zpurcon_path = download_sharepoint_file(ctx, zpurcon_relative_path, temp_dir)
        chain_pricing_path = download_sharepoint_file(ctx, chain_pricing_relative_path, temp_dir)
        if not price_book_path or not zpurcon_path or not chain_pricing_path:
            error_message = "Failed to download one or more Excel files from SharePoint."
            if not price_book_path:
                error_message += (
                    f" Could not download Price_Book_Full.xlsx from "
                    f"/teams/TDAnalysts-BBGCA/Shared Documents/{price_book_relative_path}."
                )
            if not zpurcon_path:
                error_message += (
                    f" Could not download ZPURCON.xlsx from "
                    f"/teams/TDAnalysts-BBGCA/Shared Documents/{zpurcon_relative_path}."
                )
            if not chain_pricing_path:
                error_message += (
                    f" Could not download Chain_Pricing.xlsx from "
                    f"/teams/TDAnalysts-BBGCA/Shared Documents/{chain_pricing_relative_path}."
                )
            return None, error_message
        try:
            PB = pd.read_excel(price_book_path, sheet_name="Printer Friendly")
            ZPUR = pd.read_excel(zpurcon_path)
            CHAIN = pd.read_excel(chain_pricing_path, sheet_name="Printer Friendly", header=None)
            if CHAIN.empty:
                return None, "Chain_Pricing.xlsx is empty."
            print(f"First 10 rows of Chain_Pricing.xlsx (Printer Friendly sheet):\n{CHAIN.head(10).to_string()}")
            header_row = None
            for i in range(len(CHAIN)):
                cell_value = str(CHAIN.iloc[i, 0]).strip().lower()
                if "vendor id" in cell_value:
                    header_row = i
                    break
            if header_row is None:
                return (
                    None,
                    f"Header row with 'Vendor ID' not found in column A of Chain_Pricing.xlsx. "
                    f"First 10 rows:\n{CHAIN.head(10).to_string()}"
                )
            print(f"Found header row at index {header_row}")
            CHAIN = pd.read_excel(chain_pricing_path, sheet_name="Printer Friendly", header=header_row)
            CHAIN.columns = CHAIN.columns.str.strip().str.title()
            if header_row > 0:
                CHAIN = CHAIN.iloc[header_row:].reset_index(drop=True)
            print(f"Columns in Chain_Pricing.xlsx after processing: {CHAIN.columns.tolist()}")
            print(f"Initial records in Price_Book_Full.xlsx: {len(PB)}")
            print(f"Initial records in ZPURCON.xlsx: {len(ZPUR)}")
            print(f"Initial records in Chain_Pricing.xlsx: {len(CHAIN)}")
        except FileNotFoundError as e:
            return None, f"Excel file not found: {e}"
        except Exception as e:
            return None, f"Error loading Excel files: {e}"
        if "Vendor Id" not in CHAIN.columns:
            return (
                None,
                f"Vendor ID column missing in Chain_Pricing.xlsx. Available columns: {CHAIN.columns.tolist()}"
            )
        if "Vendor Name" not in CHAIN.columns:
            return (
                None,
                f"Vendor Name column missing in Chain_Pricing.xlsx. Available columns: {CHAIN.columns.tolist()}"
            )
        # Process chain pricing data with error handling
        try:
            print("Starting chain pricing processing")
            CHAIN["Vendor Id"] = pd.Series(CHAIN["Vendor Id"], dtype="object").apply(
                lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace(".", "").isdigit() else "0"
            ).str.zfill(6)
            chain_input = CHAIN[CHAIN["Vendor Id"] == vendor_id].copy()
            print(f"Records for Vendor ID {vendor_id} in Chain_Pricing: {len(chain_input)}")
            if chain_input.empty:
                print(f"No records found for Vendor ID {vendor_id} in Chain_Pricing.")
                chain_output_df = pd.DataFrame(
                    columns=[
                        "Vendor Id",
                        "Price Group",
                        "Price Group Description",
                        "Chain Name",
                        "Start Date",
                        "End Date",
                        "List Price",
                        "Net Price",
                        "Bottle Price",
                        "Chargeback",
                        "Negotiated Cost",
                        "Avg Cost",
                        "GP2 - Negotiated Cost",
                        "GP2 - Avg Cost",
                    ]
                )
            else:
                if "Net Price" not in chain_input.columns:
                    print(
                        f"Error: Net Price column missing in Chain_Pricing.xlsx. "
                        f"Available columns: {chain_input.columns.tolist()}"
                    )
                    return None, "Net Price column missing in Chain_Pricing.xlsx."
                chain_input = chain_input.merge(
                    ZPUR[["Material", "Total", "Mov Avg 7210"]],
                    left_on="Sap Product Id",
                    right_on="Material",
                    how="left",
                ).drop(columns=["Material"], errors="ignore")
                chain_input = chain_input.rename(
                    columns={"Total": "Negotiated Cost", "Mov Avg 7210": "Avg Cost"}
                )
                print(
                    f"Non-null Net Price: {chain_input['Net Price'].notna().sum()}, "
                    f"Negotiated Cost: {chain_input['Negotiated Cost'].notna().sum()}, "
                    f"Avg Cost: {chain_input['Avg Cost'].notna().sum()}"
                )
                if "Chargeback" not in chain_input.columns:
                    chain_input["Chargeback"] = 0.0
                    print("Chargeback column missing in chain_input. Set to 0.")
                for date_col in ["Start Date", "End Date"]:
                    if date_col in chain_input.columns:
                        chain_input[date_col] = pd.to_datetime(chain_input[date_col], errors="coerce").dt.date
                if "Start Date" in chain_input.columns and "End Date" in chain_input.columns:
                    try:
                        # Ensure columns are date objects
                        chain_input["Start Date"] = pd.to_datetime(chain_input["Start Date"], errors="coerce").dt.date
                        chain_input["End Date"] = pd.to_datetime(chain_input["End Date"], errors="coerce").dt.date
                        chain_filtered = chain_input[
                            (
                                (chain_input["Start Date"].isna() | (chain_input["Start Date"] <= date_entry))
                                & (chain_input["End Date"].isna() | (chain_input["End Date"] >= date_entry))
                            )
                        ].copy()
                        print(f"Records in Chain Pricing after date filtering: {len(chain_filtered)}")
                    except Exception as e:
                        print(f"Error in Chain Pricing date filtering: {str(e)}")
                        chain_filtered = chain_input.copy()
                        print("Date filtering skipped due to error.")
                else:
                    chain_filtered = chain_input.copy()
                    print("No date filtering applied due to missing Start Date/End Date columns.")
                if "Units Per Case" in chain_filtered.columns:
                    chain_filtered["Units Per Case"] = pd.to_numeric(
                        chain_filtered["Units Per Case"], errors="coerce"
                    ).fillna(0)
                    chain_filtered["Bottle Price"] = chain_filtered.apply(
                        lambda row: row["Net Price"] / row["Units Per Case"]
                        if row["Units Per Case"] != 0
                        else pd.NA,
                        axis=1,
                    )
                else:
                    chain_filtered["Bottle Price"] = pd.NA
                    print("Units Per Case missing; Bottle Price set to NA.")
                chain_filtered = calculate_gp2_with_validation(
                    chain_filtered, skip_gp2_if_no_price=False, price_col="Net Price"
                )
                print(f"Records in Chain Pricing after GP2 calculation: {len(chain_filtered)}")
                if "Price Group" in chain_filtered.columns:
                    print(f"Price Groups in Chain Pricing: {sorted(chain_filtered['Price Group'].unique())}")
                chain_display_cols = [
                    "Vendor Id",
                    "Price Group",
                    "Price Group Description",
                    "Chain Name",
                    "Start Date",
                    "End Date",
                    "List Price",
                    "Net Price",
                    "Bottle Price",
                    "Chargeback",
                    "Negotiated Cost",
                    "Avg Cost",
                    "GP2 - Negotiated Cost",
                    "GP2 - Avg Cost",
                ]
                existing_chain_cols = [
                    col
                    for col in chain_display_cols
                    if col in chain_filtered.columns or col in ["GP2 - Negotiated Cost", "GP2 - Avg Cost"]
                ]
                chain_output_df = (
                    chain_filtered[existing_chain_cols].copy()
                    if not chain_filtered.empty
                    else pd.DataFrame(columns=chain_display_cols)
                )
                if not chain_output_df.empty:
                    chain_output_df = chain_output_df.sort_values(["Vendor Id", "Chain Name", "Price Group"])
                print(f"Final records in Chain Pricing output: {len(chain_output_df)}")
        except Exception as e:
            print(f"Error in chain pricing processing: {str(e)}")
            chain_output_df = pd.DataFrame(
                columns=[
                    "Vendor Id",
                    "Price Group",
                    "Price Group Description",
                    "Chain Name",
                    "Start Date",
                    "End Date",
                    "List Price",
                    "Net Price",
                    "Bottle Price",
                    "Chargeback",
                    "Negotiated Cost",
                    "Avg Cost",
                    "GP2 - Negotiated Cost",
                    "GP2 - Avg Cost",
                ]
            )
            print("Assigned empty chain_output_df due to error in chain pricing processing")
        cols_to_merge = [
            "Supplier",
            "Price Group #",
            "Price Group Description",
            "FOB",
            "SPA",
            "Miscellaneous",
            "Land Freight",
            "Ocean Freight",
            "Federal Tax",
            "Broker Charge",
            "Bulk Whiskey Fee",
            "Duty",
            "Tariffs Per Case",
            "Consolidate Fee",
            "Gallonage tax per case pd to Vendor",
            "Gallonage tax per case Pd to State",
            "Gallonage tax Volume based Pd to State",
            "Total",
            "Mov Avg 7210",
            "Stock in bottles",
            "Stock in Cases",
            "Mrp Controller",
        ]
        existing_cols_to_merge = [col for col in cols_to_merge if col in ZPUR.columns]
        if not existing_cols_to_merge:
            return None, "No matching columns found for merging data."
        PB_merged = PB.merge(
            ZPUR[["Material"] + existing_cols_to_merge],
            left_on="SAP Product ID",
            right_on="Material",
            how="left",
        ).drop(columns=["Material"])
        print(f"Records after merging PB and ZPUR: {len(PB_merged)}")
        if "Supplier" not in PB_merged.columns:
            return None, "Supplier column missing in merged data."
        PB_merged["Supplier"] = pd.Series(PB_merged["Supplier"], dtype="object").apply(
            lambda x: str(int(float(x))) if pd.notna(x) and str(x).replace(".", "").isdigit() else "0"
        ).str.zfill(6)
        PB_input = PB_merged[PB_merged["Supplier"] == vendor_id].copy()
        print(f"Records for Vendor ID {vendor_id}: {len(PB_input)}")
        if PB_input.empty:
            return None, f"Vendor ID {vendor_id} not found."
        vendor_name = "UnknownVendor"
        if "Vendor" in PB_input.columns and not PB_input["Vendor"].dropna().empty:
            vendor_name = str(PB_input["Vendor"].dropna().iloc[0])
        else:
            vendor_name = f"Vendor_{vendor_id}"
        vendor_name_sanitized = re.sub(r"[^\w-]", "", vendor_name).replace(" ", "_").strip("_") or "NoName"
        cols_to_drop_cogs = [
            "Vendor",
            "UPC Bottle",
            "UPC Cases",
            "UPC Sleeve",
            "Business Manager Detail",
            "Pricing Type",
            "Deal ID",
            "Deal Class",
            "Trade Channel ID",
            "Chain Name",
            "Purchase Quantity",
            "Deal Description",
            "Start Date",
            "End Date",
            "Discount",
            "Chargeback",
            "Case Price",
            "Bottle Price",
            "Units Per Case",
            "Supplier",
            "Price Group #",
            "Price Group Description_y",
            "Broker Charge",
            "Bulk Whiskey Fee",
            "Consolidate Fee",
            "Mrp Controller",
        ]
        cogs = PB_input.drop(
            columns=[col for col in cols_to_drop_cogs if col in PB_input.columns],
            errors="ignore",
        ).drop(columns=["Price Group Description_y"], errors="ignore").drop_duplicates(
            subset=["SAP Product ID"]
        )
        cogs = cogs.rename(
            columns={
                "Miscellaneous": "Misc",
                "Land Freight": "Land Frt",
                "Ocean Freight": "Ocean Frt",
                "Federal Tax": "Fed Tax",
                "Tariffs Per Case": "Tariff",
                "Gallonage tax per case pd to Vendor": "Tax pd to Ven",
                "Gallonage tax per case Pd to State": "State Tax Cs",
                "Gallonage tax Volume based Pd to State": "State Tax Vol",
                "Mov Avg 7210": "Avg Cost",
                "Stock in bottles": "Btls OH",
                "Stock in Cases": "Cases OH",
                "Price Group Description_x": "Price Group Description",
                "Total": "Negotiated Cost",
            }
        )
        if "Size" in cogs.columns:
            cogs = cogs[~cogs["Size"].str.startswith("COMBO", na=False)]
        if "List Case" in cogs.columns and "List Price" not in cogs.columns:
            cogs["List Price"] = cogs["List Case"]
        print(f"Records in COGS: {len(cogs)}")
        cols_to_drop_PW = [
            "Vendor",
            "Group Name",
            "Size",
            "Product Name",
            "SAP Product ID",
            "UPC Bottle",
            "UPC Cases",
            "UPC Sleeve",
            "Business Manager Detail",
            "Chain Name",
            "Supplier",
            "Price Group #",
            "Price Group Description_y",
            "FOB",
            "SPA",
            "Miscellaneous",
            "Land Freight",
            "Ocean Freight",
            "Federal Tax",
            "Broker Charge",
            "Bulk Whiskey Fee",
            "Duty",
            "Tariffs Per Case",
            "Consolidate Fee",
            "Gallonage tax per case pd to Vendor",
            "Gallonage tax per case Pd to State",
            "Gallonage tax Volume based Pd to State",
            "Stock in bottles",
            "Mrp Controller",
        ]
        PW = PB_input.drop(columns=[col for col in cols_to_drop_PW if col in PB_input.columns], errors="ignore")
        print(f"Records in PW before filtering: {len(PW)}")
        PW = PW.dropna(subset=["Price Group"])
        print(f"Records after dropping null Price Group: {len(PW)}")
        values_to_exclude_pricing_type = ["Volume Incentives", "Chain Pricing"]
        PW = PW[
            PW["Pricing Type"].notna()
            & (PW["Pricing Type"].astype(str).str.strip() != "")
            & (~PW["Pricing Type"].isin(values_to_exclude_pricing_type))
        ]
        print(f"Records after Pricing Type filter: {len(PW)}")
        values_to_exclude_TC = [
            "C2",
            "C3",
            "C4",
            "C5",
            "C6",
            "C7",
            "C8",
            "C9",
            "C10",
            "C11",
            "C12",
            "C13",
            "C14",
            "C15",
            "C17",
            "C18",
            "C19",
        ]
        PW = PW[~PW["Trade Channel ID"].isin(values_to_exclude_TC)]
        print(f"Records after Trade Channel ID filter: {len(PW)}")
        values_to_exclude_PQ = ["0", "0 CSE", "0 EA"]
        PW["Purchase Quantity Clean"] = PW["Purchase Quantity"].astype(str).str.strip()
        PW = PW[
            (~PW["Purchase Quantity Clean"].isin(values_to_exclude_PQ))
            & (PW["Purchase Quantity Clean"] != "")
        ]
        PW = PW.drop(columns=["Purchase Quantity Clean"])
        print(f"Records after Purchase Quantity filter: {len(PW)}")
        PW = PW.rename(
            columns={
                "Price Group Description_x": "Price Group Description",
                "Trade Channel ID": "Channel",
                "List Price": "List Case",
                "Total": "Negotiated Cost",
                "Mov Avg 7210": "Avg Cost",
                "Stock in Cases": "Cases OH",
            }
        )
        PW["Channel"] = PW["Channel"].astype(str).str.strip()
        PW["Channel"] = PW["Channel"].replace({"C1": "Retail", "C16": "OP", "": "OP", "nan": "OP"})
        for date_col in ["Start Date", "End Date"]:
            if date_col in PW.columns:
                PW[date_col] = pd.to_datetime(PW[date_col], errors="coerce").dt.date
                # Verify conversion
                if PW[date_col].dtype == "datetime64[ns]":
                    print(f"Warning: {date_col} still datetime64[ns] in PW; forcing date conversion")
                    PW[date_col] = PW[date_col].apply(lambda x: x.date() if pd.notna(x) else pd.NA)
        if "Cases OH" in PW.columns and "Avg Cost" in PW.columns:
            print(f"Processing weighted average for {len(PW)} records in PW")
            PW["Avg Cost"] = pd.to_numeric(PW["Avg Cost"], errors="coerce").fillna(0)
            PW["Cases OH"] = pd.to_numeric(PW["Cases OH"], errors="coerce").fillna(0)
            PW["FSV_Weighted_Numerator_Temp"] = PW["Avg Cost"] * PW["Cases OH"]
            PW["FSV_Weighted_Denominator_Temp"] = PW["Cases OH"]
            try:
                weighted_avg_series = PW.groupby("Price Group").agg(
                    lambda x: x["FSV_Weighted_Numerator_Temp"].sum()
                    / x["FSV_Weighted_Denominator_Temp"].sum()
                    if x["FSV_Weighted_Denominator_Temp"].sum() != 0
                    else x["Avg Cost"].iloc[0]
                    if not x["Avg Cost"].empty
                    else 0.0
                )["FSV_Weighted_Numerator_Temp"]
                print(f"weighted_avg_series type: {type(weighted_avg_series)}, length: {len(weighted_avg_series)}")
                print(f"weighted_avg_series values: {weighted_avg_series.to_list()}")
                weighted_avg_fsv_global = pd.DataFrame(
                    {
                        "Price Group": weighted_avg_series.index,
                        "Weighted_Avg_Avg_Cost_Global": weighted_avg_series.values,
                    }
                )
                PW = PW.merge(weighted_avg_fsv_global, on="Price Group", how="left")
                PW["Avg Cost"] = PW["Weighted_Avg_Avg_Cost_Global"].fillna(PW["Avg Cost"])
                PW = PW.drop(
                    columns=[
                        "FSV_Weighted_Numerator_Temp",
                        "FSV_Weighted_Denominator_Temp",
                        "Weighted_Avg_Avg_Cost_Global",
                    ],
                    errors="ignore",
                )
            except Exception as e:
                print(f"Error calculating weighted average: {str(e)}")
                PW["Avg Cost"] = PW["Avg Cost"].fillna(0)
        else:
            print("Missing 'Cases OH' or 'Avg Cost' columns; skipping weighted average calculation")
            PW["Avg Cost"] = PW["Avg Cost"].fillna(0)
        PW = PW[~PW["Price Group Description"].str.startswith("COMBO", na=False)]
        print(f"Records after COMBO filter: {len(PW)}")
        PW_deduped = improved_deduplication(PW).copy()
        print(f"Records in PW_deduped after processing: {len(PW_deduped)}")
        print(f"Price Groups in PW_deduped: {sorted(PW_deduped['Price Group'].unique())}")
        filename = (
            f"PW_{vendor_id}_{vendor_name_sanitized}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        output_path = os.path.join(output_dir, filename)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            percentage_format = "0.00%"
            light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            medium_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            side_medium = Side(style="medium")
            cogs_currency_cols = [
                "List Price",
                "FOB",
                "SPA",
                "Misc",
                "Land Frt",
                "Ocean Frt",
                "Fed Tax",
                "Duty",
                "Tariff",
                "Tax pd to Ven",
                "State Tax Cs",
                "State Tax Vol",
                "Negotiated Cost",
                "Avg Cost",
            ]
            pw_currency_cols = [
                "List Case",
                "List Bottle",
                "Discount",
                "Case Price",
                "Bottle Price",
                "Chargeback",
                "Negotiated Cost",
                "Avg Cost",
            ]
            chain_currency_cols = [
                "List Price",
                "Net Price",
                "Bottle Price",
                "Chargeback",
                "Negotiated Cost",
                "Avg Cost",
            ]
            pw_percentage_cols = ["GP2 - Negotiated Cost", "GP2 - Avg Cost"]
            cogs.to_excel(writer, index=False, sheet_name="COGS")
            worksheet = writer.sheets["COGS"]
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = light_blue_fill
                cell.font = Font(bold=True)
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 4
            worksheet.freeze_panes = "A2"
            header_map_cogs = {cell.value: cell.column for cell in worksheet[1]}
            for col_name in cogs_currency_cols:
                if col_name in header_map_cogs:
                    for row_idx in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=header_map_cogs[col_name])
                        if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.number_format = accounting_format
            if "Price Group" in cogs.columns and "Negotiated Cost" in cogs.columns:
                price_group_consistency = cogs.groupby("Price Group")["Negotiated Cost"].nunique()
                error_price_groups = price_group_consistency[price_group_consistency > 1].index.tolist()
                if error_price_groups:
                    cogs_errors_df = cogs[cogs["Price Group"].isin(error_price_groups)].sort_values(
                        ["Price Group", "SAP Product ID"]
                    )
                    if "Price Group Description_y" in cogs_errors_df.columns:
                        cogs_errors_df = cogs_errors_df.drop(columns=["Price Group Description_y"])
                    cogs_errors_df.to_excel(writer, index=False, sheet_name="Price Group Errors")
                    worksheet_errors = writer.sheets["Price Group Errors"]
                    worksheet_errors.sheet_properties.tabColor = "FF9999"
                    for col_idx in range(1, worksheet_errors.max_column + 1):
                        cell = worksheet_errors.cell(row=1, column=col_idx)
                        cell.fill = light_blue_fill
                        cell.font = Font(bold=True)
                    for column_cells in worksheet_errors.columns:
                        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                        worksheet_errors.column_dimensions[
                            get_column_letter(column_cells[0].column)
                        ].width = length + 4
                    worksheet_errors.freeze_panes = "A2"
                    header_map_errors = {cell.value: cell.column for cell in worksheet_errors[1]}
                    for col_name in cogs_currency_cols:
                        if col_name in header_map_errors:
                            for row_idx in range(2, worksheet_errors.max_row + 1):
                                cell = worksheet_errors.cell(row=row_idx, column=header_map_errors[col_name])
                                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                    cell.number_format = accounting_format
                                if col_name == "Negotiated Cost":
                                    cell.fill = light_red_fill
            gp2_filtered_df = PW_deduped.copy()
            for col in ["Units Per Case", "Case Price", "Negotiated Cost", "Chargeback", "List Case"]:
                if col in gp2_filtered_df.columns:
                    gp2_filtered_df[col] = pd.to_numeric(gp2_filtered_df[col], errors="coerce").fillna(0)
            gp2_filtered_df["List Bottle"] = gp2_filtered_df.apply(
                lambda row: row["List Case"] / row["Units Per Case"]
                if row["Units Per Case"] != 0
                else pd.NA,
                axis=1,
            )
            gp2_filtered_df["Bottle Price"] = gp2_filtered_df.apply(
                lambda row: row["Case Price"] / row["Units Per Case"]
                if row["Units Per Case"] != 0
                else pd.NA,
                axis=1,
            )
            gp2_filtered_df = calculate_gp2_with_validation(
                gp2_filtered_df, skip_gp2_if_no_price=True, price_col="Case Price"
            )
            gp2_below_threshold = gp2_filtered_df[
                (
                    (gp2_filtered_df["GP2 - Negotiated Cost"].notna())
                    & (gp2_filtered_df["GP2 - Negotiated Cost"] < gp2_threshold)
                )
                | (
                    (gp2_filtered_df["GP2 - Avg Cost"].notna())
                    & (gp2_filtered_df["GP2 - Avg Cost"] < gp2_threshold)
                )
            ].copy()
            print(f"Records in GP2 Below Threshold: {len(gp2_below_threshold)}")
            print(f"Price Groups in GP2 Below Threshold: {sorted(gp2_below_threshold['Price Group'].unique())}")
            display_cols = [
                "Price Group",
                "Price Group Description",
                "Pricing Type",
                "Deal ID",
                "Deal Class",
                "Purchase Quantity",
                "Deal Description",
                "Start Date",
                "End Date",
                "List Case",
                "Discount",
                "Chargeback",
                "Negotiated Cost",
                "Avg Cost",
                "List Bottle",
                "Case Price",
                "Bottle Price",
                "GP2 - Negotiated Cost",
                "GP2 - Avg Cost",
            ]
            existing_display_cols = [col for col in display_cols if col in gp2_below_threshold.columns]
            gp2_output_df = (
                gp2_below_threshold[existing_display_cols].copy()
                if not gp2_below_threshold.empty
                else pd.DataFrame(columns=existing_display_cols)
            )
            gp2_output_df.to_excel(writer, index=False, sheet_name="GP2 Below Threshold")
            worksheet_gp2 = writer.sheets["GP2 Below Threshold"]
            worksheet_gp2.sheet_properties.tabColor = "FF9999"
            for col_idx in range(1, worksheet_gp2.max_column + 1):
                cell = worksheet_gp2.cell(row=1, column=col_idx)
                cell.fill = light_blue_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            for column_cells in worksheet_gp2.columns:
                column_letter = get_column_letter(column_cells[0].column)
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                if column_letter in ["M", "N", "O", "P", "Q"]:
                    worksheet_gp2.column_dimensions[column_letter].width = max(length + 6, 15)
                else:
                    worksheet_gp2.column_dimensions[column_letter].width = length + 4
            worksheet_gp2.freeze_panes = "A2"
            if not gp2_output_df.empty:
                header_map_gp2 = {cell.value: cell.column for cell in worksheet_gp2[1]}
                for col_name in pw_currency_cols:
                    if col_name in header_map_gp2:
                        for row_idx in range(2, worksheet_gp2.max_row + 1):
                            cell = worksheet_gp2.cell(row=row_idx, column=header_map_gp2[col_name])
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = accounting_format
                for col_name in pw_percentage_cols:
                    if col_name in header_map_gp2:
                        for row_idx in range(2, worksheet_gp2.max_row + 1):
                            cell = worksheet_gp2.cell(row=row_idx, column=header_map_gp2[col_name])
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = percentage_format
                                if cell.value < gp2_threshold:
                                    cell.fill = light_red_fill
                for date_col in ["Start Date", "End Date"]:
                    if date_col in header_map_gp2:
                        for row_idx in range(2, worksheet_gp2.max_row + 1):
                            cell = worksheet_gp2.cell(row=row_idx, column=header_map_gp2[date_col])
                            if isinstance(cell.value, (datetime, pd.Timestamp)) and not pd.isna(cell.value):
                                cell.number_format = "MM/DD/YYYY"
            else:
                worksheet_gp2.cell(
                    row=2, column=1, value=f"No records found with GP2 margins below {gp2_threshold:.1%}"
                )
                worksheet_gp2.cell(row=2, column=1).font = Font(italic=True, color="666666")
            deal_id_df = PW_deduped.copy()
            for col in ["Units Per Case", "Case Price", "Negotiated Cost", "Chargeback", "List Case"]:
                if col in deal_id_df.columns:
                    deal_id_df[col] = pd.to_numeric(deal_id_df[col], errors="coerce").fillna(0)
            deal_id_df["List Bottle"] = deal_id_df.apply(
                lambda row: row["List Case"] / row["Units Per Case"]
                if row["Units Per Case"] != 0
                else pd.NA,
                axis=1,
            )
            deal_id_df["Bottle Price"] = deal_id_df.apply(
                lambda row: row["Case Price"] / row["Units Per Case"]
                if row["Units Per Case"] != 0
                else pd.NA,
                axis=1,
            )
            deal_id_df = calculate_gp2_with_validation(
                deal_id_df, skip_gp2_if_no_price=True, price_col="Case Price"
            )
            if "Start Date" in deal_id_df.columns and "End Date" in deal_id_df.columns:
                try:
                    # Ensure columns are date objects
                    deal_id_df["Start Date"] = pd.to_datetime(deal_id_df["Start Date"], errors="coerce").dt.date
                    deal_id_df["End Date"] = pd.to_datetime(deal_id_df["End Date"], errors="coerce").dt.date
                    # Verify conversion
                    for date_col in ["Start Date", "End Date"]:
                        if deal_id_df[date_col].dtype == "datetime64[ns]":
                            print(f"Warning: {date_col} still datetime64[ns] in deal_id_df; forcing date conversion")
                            deal_id_df[date_col] = deal_id_df[date_col].apply(
                                lambda x: x.date() if pd.notna(x) else pd.NA
                            )
                    deal_id_df = deal_id_df[
                        (
                            (deal_id_df["Start Date"].isna())
                            | (deal_id_df["End Date"].isna())
                            | (deal_id_df["Start Date"] <= date_entry)
                        )
                    ]
                    print(f"Records in Pricing by Deal ID after relaxed date filtering: {len(deal_id_df)}")
                    print(f"Price Groups in Pricing by Deal ID: {sorted(deal_id_df['Price Group'].unique())}")
                except Exception as e:
                    print(f"Error in Pricing by Deal ID date filtering: {str(e)}")
                    print("Skipping date filtering due to error.")
            deal_id_display_cols = [
                "Price Group",
                "Price Group Description",
                "Pricing Type",
                "Deal ID",
                "Deal Class",
                "Purchase Quantity",
                "Deal Description",
                "Start Date",
                "End Date",
                "List Case",
                "Discount",
                "Chargeback",
                "Negotiated Cost",
                "Avg Cost",
                "List Bottle",
                "Case Price",
                "Bottle Price",
                "GP2 - Negotiated Cost",
                "GP2 - Avg Cost",
            ]
            existing_deal_id_cols = [col for col in deal_id_display_cols if col in deal_id_df.columns]
            deal_id_output_df = (
                deal_id_df[existing_deal_id_cols].copy()
                if not deal_id_df.empty
                else pd.DataFrame(columns=existing_deal_id_cols)
            )
            if not deal_id_output_df.empty:
                deal_id_output_df = deal_id_output_df.sort_values(["Deal ID", "Deal Description"])
            print(f"Final records in Pricing by Deal ID output: {len(deal_id_output_df)}")
            deal_id_output_df.to_excel(writer, index=False, sheet_name="Pricing by Deal ID")
            worksheet_deal_id = writer.sheets["Pricing by Deal ID"]
            for col_idx in range(1, worksheet_deal_id.max_column + 1):
                cell = worksheet_deal_id.cell(row=1, column=col_idx)
                cell.fill = light_blue_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            for column_cells in worksheet_deal_id.columns:
                column_letter = get_column_letter(column_cells[0].column)
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                if column_letter in ["K", "L", "M", "N", "O", "R", "S"]:
                    worksheet_deal_id.column_dimensions[column_letter].width = max(length + 6, 15)
                else:
                    worksheet_deal_id.column_dimensions[column_letter].width = length + 4
            worksheet_deal_id.freeze_panes = "A2"
            if not deal_id_output_df.empty:
                header_map_deal_id = {cell.value: cell.column for cell in worksheet_deal_id[1]}
                for col_name in pw_currency_cols:
                    if col_name in header_map_deal_id:
                        for row_idx in range(2, worksheet_deal_id.max_row + 1):
                            cell = worksheet_deal_id.cell(row=row_idx, column=header_map_deal_id[col_name])
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = accounting_format
                for col_name in pw_percentage_cols:
                    if col_name in header_map_deal_id:
                        for row_idx in range(2, worksheet_deal_id.max_row + 1):
                            cell = worksheet_deal_id.cell(row=row_idx, column=header_map_deal_id[col_name])
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = percentage_format
                                if cell.value < gp2_threshold:
                                    cell.fill = light_red_fill
                for date_col in ["Start Date", "End Date"]:
                    if date_col in header_map_deal_id:
                        for row_idx in range(2, worksheet_deal_id.max_row + 1):
                            cell = worksheet_deal_id.cell(row=row_idx, column=header_map_deal_id[date_col])
                            if isinstance(cell.value, (datetime, pd.Timestamp)) and not pd.isna(cell.value):
                                cell.number_format = "MM/DD/YYYY"
            else:
                worksheet_deal_id.cell(
                    row=2,
                    column=1,
                    value=f"No pricing records found for vendor {vendor_id} on {date_entry}",
                )
                worksheet_deal_id.cell(row=2, column=1).font = Font(italic=True, color="666666")
            chain_output_df.to_excel(writer, index=False, sheet_name="Chain Pricing")
            worksheet_chain = writer.sheets["Chain Pricing"]
            for col_idx in range(1, worksheet_chain.max_column + 1):
                cell = worksheet_chain.cell(row=1, column=col_idx)
                cell.fill = light_blue_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            for column_cells in worksheet_chain.columns:
                column_letter = get_column_letter(column_cells[0].column)
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet_chain.column_dimensions[column_letter].width = length + 4
            worksheet_chain.freeze_panes = "A2"
            if not chain_output_df.empty:
                header_map_chain = {cell.value: cell.column for cell in worksheet_chain[1]}
                for col_name in chain_currency_cols:
                    if col_name in header_map_chain:
                        for row_idx in range(2, worksheet_chain.max_row + 1):
                            cell = worksheet_chain.cell(row=row_idx, column=header_map_chain[col_name])
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = accounting_format
                for col_name in pw_percentage_cols:
                    if col_name in header_map_chain:
                        for row_idx in range(2, worksheet_chain.max_row + 1):
                            cell = worksheet_chain.cell(row=row_idx, column=header_map_chain[col_name])
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = percentage_format
                                if cell.value < gp2_threshold:
                                    cell.fill = light_red_fill
                for date_col in ["Start Date", "End Date"]:
                    if date_col in header_map_chain:
                        for row_idx in range(2, worksheet_chain.max_row + 1):
                            cell = worksheet_chain.cell(row=row_idx, column=header_map_chain[date_col])
                            if isinstance(cell.value, (datetime, pd.Timestamp)) and not pd.isna(cell.value):
                                cell.number_format = "MM/DD/YYYY"
            else:
                worksheet_chain.cell(
                    row=2,
                    column=1,
                    value=f"No chain pricing records found for vendor {vendor_id} on {date_entry}",
                )
                worksheet_chain.cell(row=2, column=1).font = Font(italic=True, color="666666")
            if "Brand" in PW_deduped.columns:
                print(f"Brand column found. Unique brands: {PW_deduped['Brand'].dropna().unique()}")
                if PW_deduped["Brand"].dropna().empty:
                    print("Warning: No non-NaN Brand values found in PW_deduped.")
                    pd.DataFrame({"Message": ["No valid Brand data found"]}).to_excel(
                        writer, index=False, sheet_name="No_Brands"
                    )
                    worksheet_no_brands = writer.sheets["No_Brands"]
                    worksheet_no_brands.cell(row=1, column=1).font = Font(italic=True, color="666666")
                else:
                    for brand in PW_deduped["Brand"].dropna().unique():
                        print(f"Processing brand: {brand}")
                        brand_df = PW_deduped[PW_deduped["Brand"] == brand].copy()
                        if brand_df.empty:
                            print(f"Warning: No data for brand {brand} after filtering.")
                            continue
                        # Ensure numeric columns
                        for col in ["Units Per Case", "Case Price", "Negotiated Cost", "Chargeback", "List Case"]:
                            if col in brand_df.columns:
                                brand_df[col] = pd.to_numeric(brand_df[col], errors="coerce").fillna(0)
                        # Calculate derived columns
                        brand_df["List Bottle"] = brand_df.apply(
                            lambda row: row["List Case"] / row["Units Per Case"]
                            if row["Units Per Case"] != 0
                            else pd.NA,
                            axis=1,
                        )
                        brand_df["Bottle Price"] = brand_df.apply(
                            lambda row: row["Case Price"] / row["Units Per Case"]
                            if row["Units Per Case"] != 0
                            else pd.NA,
                            axis=1,
                        )
                        # Calculate GP2 fields
                        brand_df = calculate_gp2_with_validation(
                            brand_df, skip_gp2_if_no_price=True, price_col="Case Price"
                        )
                        # Create Pivot Key
                        pivot_key_cols = ["Channel", "Pricing Type", "Deal Class", "Purchase Quantity"]
                        existing_pivot_cols = [col for col in pivot_key_cols if col in brand_df.columns]
                        print(f"Brand: {brand}, Available pivot columns: {existing_pivot_cols}")
                        if len(existing_pivot_cols) == 4:
                            brand_df["Channel"] = brand_df["Channel"].astype(str).str.strip()
                            brand_df["Channel"] = brand_df["Channel"].replace(
                                {"C1": "Retail", "C16": "OP", "": "OP", "nan": "OP"}
                            )
                            brand_df["Pivot Key"] = brand_df[existing_pivot_cols].astype(str).agg(" | ".join, axis=1)
                            print(f"Brand: {brand}, Unique Pivot Keys: {brand_df['Pivot Key'].unique()}")
                            brand_df["Pivot Key Valid"] = brand_df["Pivot Key"].apply(
                                lambda x: len(x.split(" | ")) == 4 and x.split(" | ")[0] in ["Retail", "OP"]
                            )
                            if not brand_df["Pivot Key Valid"].all():
                                print(f"Warning: Invalid Pivot Keys found for brand {brand}:")
                                print(brand_df[~brand_df["Pivot Key Valid"]]["Pivot Key"].unique())
                            # Define id_vars and value_vars
                            id_vars = [
                                col
                                for col in [
                                    "Vendor ID",
                                    "Brand",
                                    "Price Group",
                                    "Price Group Description",
                                    "Units Per Case",
                                    "Pivot Key",
                                ]
                                if col in brand_df.columns
                            ]
                            value_vars = [
                                col
                                for col in [
                                    "Negotiated Cost",
                                    "Avg Cost",
                                    "List Case",
                                    "List Bottle",
                                    "Discount",
                                    "Case Price",
                                    "Bottle Price",
                                    "Chargeback",
                                    "GP2 - Negotiated Cost",
                                    "GP2 - Avg Cost",
                                    "Start Date",
                                    "End Date",
                                    "Deal ID",
                                    "Deal Description",
                                ]
                                if col in brand_df.columns
                            ]
                            try:
                                # Melt and pivot
                                brand_melted = brand_df.melt(
                                    id_vars=id_vars,
                                    value_vars=value_vars,
                                    var_name="Product Cost Breakdown",
                                    value_name="Value",
                                )
                                brand_pivot = brand_melted.pivot_table(
                                    index=[
                                        col
                                        for col in [
                                            "Brand",
                                            "Price Group",
                                            "Price Group Description",
                                            "Units Per Case",
                                            "Product Cost Breakdown",
                                        ]
                                        if col in brand_melted.columns
                                    ],
                                    columns="Pivot Key",
                                    values="Value",
                                    aggfunc="first",
                                ).reset_index().drop(columns=["Brand"], errors="ignore")
                                # Define desired order for Product Cost Breakdown
                                desired_order = [
                                    col for col in value_vars if col in brand_pivot["Product Cost Breakdown"].values
                                ]
                                if "Product Cost Breakdown" in brand_pivot.columns:
                                    brand_pivot["Product Cost Breakdown"] = pd.Categorical(
                                        brand_pivot["Product Cost Breakdown"],
                                        categories=desired_order,
                                        ordered=True,
                                    )
                                    brand_pivot = brand_pivot.sort_values(
                                        ["Price Group", "Price Group Description", "Units Per Case", "Product Cost Breakdown"]
                                    )
                                    # Blank out repeating values
                                    group_cols = ["Price Group", "Price Group Description", "Units Per Case"]
                                    if group_cols[0] in brand_pivot.columns:
                                        block_change = brand_pivot[group_cols].ne(
                                            brand_pivot[group_cols].shift()
                                        ).any(axis=1)
                                        for col in group_cols:
                                            if col in brand_pivot.columns:
                                                brand_pivot[col] = brand_pivot[col].where(block_change, "")
                                # Sort Pivot Key columns
                                pivot_cols = [
                                    col
                                    for col in brand_pivot.columns
                                    if col
                                    not in ["Price Group", "Price Group Description", "Units Per Case", "Product Cost Breakdown"]
                                ]
                                print(f"Brand: {brand}, Pivot Columns Before Sorting: {pivot_cols}")
                                def pivot_key_sort_key(key):
                                    try:
                                        parts = key.split(" | ")
                                        channel = parts[0] if len(parts) > 0 else ""
                                        pricing_type = parts[1] if len(parts) > 1 else ""
                                        deal_class = parts[2] if len(parts) > 2 else ""
                                        purchase_qty = parts[3] if len(parts) > 3 else ""
                                        # Channel order
                                        channel_order = {"Retail": 0, "OP": 1}
                                        channel_val = channel_order.get(channel, 99)
                                        # Pricing Type order
                                        pricing_order = {"Level Pricing": 0, "Deal Pricing": 1}
                                        pricing_val = pricing_order.get(pricing_type, 99)
                                        # Deal Class order
                                        deal_class_order = [
                                            "Level Pricing",
                                            "EVD – Straight Discount",
                                            "Close – Straight Discount",
                                            "Promo – Straight Discount",
                                            "EVD- Special Price Goods",
                                            "Promo- Special Price Goods",
                                            "Inventory Reduction – Straight Discount",
                                            "Inventory Reduction- Special Price Goods",
                                        ]
                                        deal_class_val = (
                                            deal_class_order.index(deal_class)
                                            if deal_class in deal_class_order
                                            else 99
                                        )
                                        # Purchase Quantity numeric value
                                        qty_match = re.match(r"(\d+)", purchase_qty)
                                        qty_val = int(qty_match.group(1)) if qty_match else 9999
                                        return (channel_val, pricing_val, deal_class_val, qty_val)
                                    except Exception:
                                        return (99, 99, 99, 9999)
                                sorted_pivot_cols = sorted(pivot_cols, key=pivot_key_sort_key)
                                print(f"Brand: {brand}, Pivot Columns After Sorting: {sorted_pivot_cols}")
                                brand_pivot = brand_pivot[
                                    ["Price Group", "Price Group Description", "Units Per Case", "Product Cost Breakdown"]
                                    + sorted_pivot_cols
                                ]
                                # Write pivot table to Excel
                                safe_brand_name = str(brand)[:31]
                                brand_pivot.to_excel(writer, index=False, sheet_name=safe_brand_name)
                                print(f"Successfully wrote pivot table for brand {brand}")
                                # Formatting
                                worksheet_brand = writer.sheets[safe_brand_name]
                                # Insert 3 rows for stacked headers
                                worksheet_brand.insert_rows(1, amount=3)
                                # Apply light blue fill and bold font to header rows (1-4)
                                for r_idx in range(1, 5):
                                    for c_idx in range(1, worksheet_brand.max_column + 1):
                                        cell = worksheet_brand.cell(row=r_idx, column=c_idx)
                                        cell.fill = light_blue_fill
                                        cell.font = Font(bold=True)
                                # Define header map for row 4
                                final_row_4_headers_map = {
                                    "Price Group": "Price Group",
                                    "Price Group Description": "Price Group Description",
                                    "Units Per Case": "PK",
                                    "Product Cost Breakdown": "Pricing Details",
                                }
                                # Set headers in rows 1-4
                                for col_idx in range(1, worksheet_brand.max_column + 1):
                                    cell_in_row4 = worksheet_brand.cell(row=4, column=col_idx)
                                    original_header = cell_in_row4.value
                                    if original_header in final_row_4_headers_map:
                                        target_cell = worksheet_brand.cell(row=4, column=col_idx)
                                        target_cell.value = final_row_4_headers_map[original_header]
                                        target_cell.alignment = Alignment(horizontal="center")
                                        for row_offset in range(3):
                                            worksheet_brand.cell(row=1 + row_offset, column=col_idx, value="")
                                    elif original_header and isinstance(original_header, str) and " | " in original_header:
                                        parts = original_header.split(" | ")
                                        while len(parts) < 4:
                                            parts.append("")
                                        header_parts_to_write = [parts[0], parts[1], parts[2], parts[3]]
                                        for row_offset, part_value in enumerate(header_parts_to_write):
                                            header_cell = worksheet_brand.cell(
                                                row=1 + row_offset, column=col_idx, value=part_value
                                            )
                                            header_cell.alignment = Alignment(horizontal="center")
                                    else:
                                        target_cell = worksheet_brand.cell(row=4, column=col_idx)
                                        target_cell.alignment = Alignment(horizontal="center")
                                        for row_offset in range(3):
                                            worksheet_brand.cell(row=1 + row_offset, column=col_idx, value="")
                                # Freeze panes at E5
                                worksheet_brand.freeze_panes = "E5"
                                # Apply borders
                                pcb_idx = brand_pivot.columns.get_loc("Product Cost Breakdown") + 1
                                pivot_key_excel_col_indices = []
                                start_col_for_pivot_keys = pcb_idx + 1
                                for col_idx_openpyxl in range(start_col_for_pivot_keys, worksheet_brand.max_column + 1):
                                    pivot_key_excel_col_indices.append(col_idx_openpyxl)
                                rows_for_bottom_border = []
                                for r_idx_check in range(5, worksheet_brand.max_row + 1):
                                    pcb_cell_value = worksheet_brand.cell(row=r_idx_check, column=pcb_idx).value
                                    if pcb_cell_value == "Deal Description":
                                        rows_for_bottom_border.append(r_idx_check)
                                for c_idx in range(1, worksheet_brand.max_column + 1):
                                    cell = worksheet_brand.cell(row=4, column=c_idx)
                                    current_border = cell.border if cell.border else Border()
                                    cell.border = Border(
                                        left=current_border.left,
                                        right=current_border.right,
                                        top=current_border.top,
                                        bottom=side_medium,
                                    )
                                for r_idx in range(1, worksheet_brand.max_row + 1):
                                    for c_idx in range(1, worksheet_brand.max_column + 1):
                                        cell = worksheet_brand.cell(row=r_idx, column=c_idx)
                                        current_border = cell.border if cell.border else Border()
                                        new_left = current_border.left
                                        new_right = current_border.right
                                        new_top = current_border.top
                                        new_bottom = current_border.bottom
                                        if r_idx in rows_for_bottom_border:
                                            new_bottom = side_medium
                                        if c_idx == pcb_idx:
                                            new_right = side_medium
                                        if c_idx in pivot_key_excel_col_indices:
                                            new_left = side_medium
                                            if c_idx == pivot_key_excel_col_indices[-1] or (c_idx + 1) not in pivot_key_excel_col_indices:
                                                new_right = side_medium
                                        cell.border = Border(
                                            left=new_left, right=new_right, top=new_top, bottom=new_bottom
                                        )
                                # Auto-size columns
                                for column_cells in worksheet_brand.columns:
                                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                                    column_letter = get_column_letter(column_cells[0].column)
                                    worksheet_brand.column_dimensions[column_letter].width = length + 4
                                # Center align Price Group, Price Group Description, Units Per Case
                                center_alignment = Alignment(horizontal="center")
                                cols_to_center_data = ["Price Group", "Price Group Description", "Units Per Case"]
                                for col_name in cols_to_center_data:
                                    if col_name in brand_pivot.columns:
                                        col_idx = brand_pivot.columns.get_loc(col_name) + 1
                                        for row in worksheet_brand.iter_rows(min_col=col_idx, max_col=col_idx, min_row=5):
                                            for cell in row:
                                                cell.alignment = center_alignment
                                # Right align Product Cost Breakdown and Deal Description values
                                right_alignment = Alignment(horizontal="right")
                                for row_idx in range(5, worksheet_brand.max_row + 1):
                                    cell_pcb = worksheet_brand.cell(row=row_idx, column=pcb_idx)
                                    cell_pcb.alignment = right_alignment
                                    if cell_pcb.value == "Deal Description":
                                        for col_idx in range(pcb_idx + 1, worksheet_brand.max_column + 1):
                                            value_cell = worksheet_brand.cell(row=row_idx, column=col_idx)
                                            value_cell.alignment = right_alignment
                                # Apply number formatting
                                header_map_brand = {cell.value: cell.column for cell in worksheet_brand[4]}
                                for col_name in pw_currency_cols:
                                    if col_name in header_map_brand:
                                        for row_idx in range(5, worksheet_brand.max_row + 1):
                                            cell = worksheet_brand.cell(row=row_idx, column=header_map_brand[col_name])
                                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                                cell.number_format = accounting_format
                                # Apply percentage format to GP2 fields and conditional highlighting
                                data_start_row = 5  # Data starts after header rows 1-4
                                for row_idx in range(data_start_row, worksheet_brand.max_row + 1):
                                    pcb_value = worksheet_brand.cell(row=row_idx, column=pcb_idx).value
                                    if pcb_value in pw_percentage_cols:  # Check if it's one of the GP2 rows
                                        for col_idx in range(pcb_idx + 1, worksheet_brand.max_column + 1):  # Iterate through value columns
                                            value_cell = worksheet_brand.cell(row=row_idx, column=col_idx)
                                            if isinstance(value_cell.value, (int, float)) and not pd.isna(value_cell.value):
                                                value_cell.number_format = percentage_format
                                                if value_cell.value < gp2_threshold:
                                                    value_cell.fill = light_red_fill
                                for date_col in ["Start Date", "End Date"]:
                                    if date_col in header_map_brand:
                                        for row_idx in range(5, worksheet_brand.max_row + 1):
                                            cell = worksheet_brand.cell(row=row_idx, column=header_map_brand[date_col])
                                            if isinstance(cell.value, (datetime, pd.Timestamp)) and not pd.isna(cell.value):
                                                cell.number_format = "MM/DD/YYYY"
                            except Exception as e:
                                print(f"Error creating pivot table for brand {brand}: {str(e)}")
                                brand_cols = [col for col in display_cols if col in brand_df.columns]
                                brand_simple = brand_df[brand_cols].copy()
                                safe_brand_name = str(brand)[:31]
                                brand_simple.to_excel(writer, index=False, sheet_name=safe_brand_name)
                                print(f"Fallback: Wrote simple table for brand {brand}")
                                continue
                        else:
                            print(f"Brand: {brand}, Insufficient columns for pivot table: {existing_pivot_cols}")
                            brand_cols = [col for col in display_cols if col in brand_df.columns]
                            brand_simple = brand_df[brand_cols].copy()
                            safe_brand_name = str(brand)[:31]
                            brand_simple.to_excel(writer, index=False, sheet_name=safe_brand_name)
                            print(f"Wrote simple table for brand {brand}")
            else:
                print("No Brand column found in PW_deduped.")
                pd.DataFrame({"Message": ["No Brand column available"]}).to_excel(
                    writer, index=False, sheet_name="No_Brands"
                )
                worksheet_no_brands = writer.sheets["No_Brands"]
                worksheet_no_brands.cell(row=1, column=1).font = Font(italic=True, color="666666")
        return filename, None
    finally:
        try:
            shutil.rmtree(temp_dir)
            print(f"Cleaned up temporary directory: {temp_dir}")
        except Exception as e:
            print(f"Error cleaning up temporary directory {temp_dir}: {str(e)}")

@app.route("/", methods=["GET", "POST"])
def index():
    """Handle the main page and form submission."""
    if request.method == "POST":
        vendor_id = request.form.get("vendor_id", "").strip()
        gp2_threshold = request.form.get("gp2_threshold", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        date_entry = request.form.get("date_entry", "").strip()
        if not email or not password:
            flash("Email and password are required.", "error")
            return redirect(url_for("index"))
        if not (vendor_id.isdigit() and len(vendor_id) == 6 and vendor_id.startswith("3")):
            flash("Invalid Vendor ID.", "error")
            return redirect(url_for("index"))
        try:
            gp2_threshold_val = float(gp2_threshold)
            if not (0 <= gp2_threshold_val <= 1):
                raise ValueError()
        except ValueError:
            flash("Invalid GP2 threshold.", "error")
            return redirect(url_for("index"))
        filename, error = process_data(vendor_id, gp2_threshold_val, email, password, date_entry)
        if error:
            flash(error, "error")
            return redirect(url_for("index"))
        flash(f"Processing complete! Download: {filename}", "success")
        return redirect(url_for("download_file", filename=filename))
    return render_template("index2.html")

@app.route("/api/process", methods=["POST"])
def api_process():
    """Handle API requests for processing data."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "No JSON data received"}), 400
        vendor_id = str(data.get("vendor_id", "")).strip()
        gp2_threshold = str(data.get("gp2_threshold", "")).strip()
        email = str(data.get("email", "")).strip()
        password = str(data.get("password", "")).strip()
        date_entry = str(data.get("date_entry", "")).strip()
        if not email or not password:
            return jsonify({"success": False, "message": "Email and password are required."}), 400
        if not (vendor_id.isdigit() and len(vendor_id) == 6 and vendor_id.startswith("3")):
            return jsonify({"success": False, "message": "Invalid Vendor ID."}), 400
        try:
            gp2_threshold_val = float(gp2_threshold)
            if not (0 <= gp2_threshold_val <= 1):
                raise ValueError()
        except ValueError:
            return jsonify({"success": False, "message": "Invalid GP2 threshold."}), 400
        filename, error = process_data(vendor_id, gp2_threshold_val, email, password, date_entry)
        if error:
            return jsonify({"success": False, "message": error}), 500
        return jsonify(
            {
                "success": True,
                "message": f"File generated: {filename}",
                "download_url": url_for("download_file", filename=filename, _external=True),
            }
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500

@app.route("/download/<filename>")
def download_file(filename):
    """Serve the generated Excel file for download."""
    path = os.path.join(output_dir, filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name=filename)
    flash("File not found.", "error")
    return redirect(url_for("index"))

def open_browser():
    """Open the default web browser to the app's URL."""
    webbrowser.open_new("http://127.0.0.1:5000")

if __name__ == "__main__":
    threading.Timer(1.0, open_browser).start()
    app.run(debug=False, host="127.0.0.1", port=5000)